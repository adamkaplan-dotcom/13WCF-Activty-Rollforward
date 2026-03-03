"""
USDx Balances Weekly Update Script

This script reads weekly balance data from a source file and updates a master tracking file.

Usage:
    python update_balances.py <source_file> <target_file> [output_file]

Example:
    python update_balances.py "Weekly_Forecast_2026-02-20.xlsx" "Master_Balances.xlsx"
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict


def read_source_file(source_path):
    """
    Step 1: Read the USDx Balances tab from the source file.

    Returns:
        dict: {
            'balance_date': datetime,
            'balances': {ref_number: balance_value, ...}
        }
    """
    print(f"\n{'='*60}")
    print("STEP 1: Reading source file")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(source_path, data_only=True)

    # Find the USDx Balances sheet
    sheet = None
    for sheet_name in wb.sheetnames:
        if "USDx Balances" in sheet_name or "USDx" in sheet_name:
            sheet = wb[sheet_name]
            print(f"✓ Found sheet: {sheet_name}")
            break

    if sheet is None:
        raise ValueError("Could not find 'USDx Balances' sheet in source file")

    # Read header row (row 12)
    header_row = 12
    balance_date = None
    current_week_col = None

    # Find "Current Week" column (typically column K = 11)
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(header_row, col_idx).value
        if cell_value and isinstance(cell_value, datetime):
            balance_date = cell_value
            current_week_col = col_idx
            print(f"✓ Found balance date: {balance_date.strftime('%Y-%m-%d')} in column {get_column_letter(col_idx)}")
            break
        elif cell_value and "current" in str(cell_value).lower():
            # If it's a text header, check the cell below for date
            date_cell = sheet.cell(header_row + 1, col_idx).value
            if isinstance(date_cell, datetime):
                balance_date = date_cell
                current_week_col = col_idx
                print(f"✓ Found balance date: {balance_date.strftime('%Y-%m-%d')} in column {get_column_letter(col_idx)}")
                break

    if not balance_date or not current_week_col:
        # Fallback: assume column K (11)
        current_week_col = 11
        balance_date = sheet.cell(header_row, current_week_col).value
        print(f"⚠ Using fallback column K for current week")

    # Read data rows (ref number in col B, balance in current_week_col)
    balances = {}
    ref_col = 2  # Column B
    skipped_rows = []
    data_start_row = header_row + 1

    print(f"\nReading balances from row {data_start_row} onwards...")

    for row_idx in range(data_start_row, sheet.max_row + 1):
        ref_value = sheet.cell(row_idx, ref_col).value

        # Skip empty rows
        if ref_value is None:
            continue

        # Skip section headers (text values that are descriptive)
        ref_str = str(ref_value).strip()
        if any(keyword in ref_str.lower() for keyword in ['bullish', 'coindesk', 'fiat', 'balance', 'total']):
            skipped_rows.append((row_idx, ref_str))
            continue

        # Get balance value
        balance_value = sheet.cell(row_idx, current_week_col).value

        # Convert to float if possible
        if balance_value is not None:
            try:
                balance_value = float(balance_value)
            except (ValueError, TypeError):
                balance_value = 0.0
        else:
            balance_value = 0.0

        balances[ref_str] = balance_value

    print(f"✓ Extracted {len(balances)} balance records")
    if skipped_rows:
        print(f"✓ Skipped {len(skipped_rows)} section header rows")

    wb.close()

    return {
        'balance_date': balance_date,
        'balances': balances
    }


def read_target_file(target_path):
    """
    Step 2: Read the target/master file.

    Returns:
        dict: {
            'workbook': workbook object,
            'sheet': sheet object,
            'ref_col': column index for ref numbers,
            'last_data_col': last populated column index,
            'ref_rows': {ref_number: row_index, ...},
            'existing_dates': [list of date headers]
        }
    """
    print(f"\n{'='*60}")
    print("STEP 2: Reading target file")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(target_path)
    sheet = wb.active

    print(f"✓ Loaded workbook: {target_path}")
    print(f"✓ Active sheet: {sheet.title}")

    # Identify ref number column (assume column B = 2)
    ref_col = 2

    # Find the last populated column
    last_data_col = 1
    header_row = 1  # Assuming headers are in row 1

    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(header_row, col_idx).value
        if cell_value is not None:
            last_data_col = col_idx

    print(f"✓ Reference column: {get_column_letter(ref_col)}")
    print(f"✓ Last populated column: {get_column_letter(last_data_col)}")

    # Read existing date headers
    existing_dates = []
    for col_idx in range(3, last_data_col + 1):  # Start from col C onwards
        header_val = sheet.cell(header_row, col_idx).value
        if isinstance(header_val, datetime):
            existing_dates.append(header_val)
        elif header_val:
            try:
                # Try to parse as date string
                date_obj = datetime.strptime(str(header_val), '%Y-%m-%d')
                existing_dates.append(date_obj)
            except:
                pass

    print(f"✓ Found {len(existing_dates)} existing date columns")
    if existing_dates:
        print(f"  Most recent: {max(existing_dates).strftime('%Y-%m-%d')}")

    # Build ref number to row mapping
    ref_rows = {}
    data_start_row = 2  # Assuming data starts at row 2

    for row_idx in range(data_start_row, sheet.max_row + 1):
        ref_value = sheet.cell(row_idx, ref_col).value
        if ref_value:
            ref_rows[str(ref_value).strip()] = row_idx

    print(f"✓ Found {len(ref_rows)} reference numbers in target file")

    return {
        'workbook': wb,
        'sheet': sheet,
        'ref_col': ref_col,
        'last_data_col': last_data_col,
        'ref_rows': ref_rows,
        'existing_dates': existing_dates,
        'header_row': header_row
    }


def update_target_file(source_data, target_data):
    """
    Step 3: Match and insert new column with balances.

    Returns:
        dict: Statistics about the update
    """
    print(f"\n{'='*60}")
    print("STEP 3: Updating target file with new data")
    print(f"{'='*60}")

    sheet = target_data['sheet']
    balance_date = source_data['balance_date']
    balances = source_data['balances']
    existing_dates = target_data['existing_dates']

    # Check if this date already exists
    if balance_date in existing_dates:
        print(f"⚠ WARNING: Date {balance_date.strftime('%Y-%m-%d')} already exists in target file!")
        response = input("Continue and overwrite? (y/n): ")
        if response.lower() != 'y':
            print("Aborted.")
            return None

    # Add new column after the last data column
    new_col = target_data['last_data_col'] + 1
    header_row = target_data['header_row']

    # Write header
    sheet.cell(header_row, new_col).value = balance_date
    print(f"✓ Added new column {get_column_letter(new_col)} with date: {balance_date.strftime('%Y-%m-%d')}")

    # Match and write balances
    matched = 0
    missing = 0
    missing_refs = []

    for ref_num, row_idx in target_data['ref_rows'].items():
        if ref_num in balances:
            sheet.cell(row_idx, new_col).value = balances[ref_num]
            matched += 1
        else:
            sheet.cell(row_idx, new_col).value = 0.0
            missing += 1
            missing_refs.append(ref_num)

    # Find new refs in source that don't exist in target
    new_refs = []
    for ref_num in balances.keys():
        if ref_num not in target_data['ref_rows']:
            new_refs.append((ref_num, balances[ref_num]))

    print(f"\n✓ Matched and updated: {matched} records")
    if missing > 0:
        print(f"⚠ Missing in source: {missing} records (set to 0.0)")
    if new_refs:
        print(f"⚠ New refs found in source: {len(new_refs)} (not in target)")
        print(f"  These refs need to be added manually:")
        for ref, bal in new_refs[:5]:  # Show first 5
            print(f"    - {ref}: {bal}")
        if len(new_refs) > 5:
            print(f"    ... and {len(new_refs) - 5} more")

    return {
        'matched': matched,
        'missing': missing,
        'missing_refs': missing_refs,
        'new_refs': new_refs
    }


def save_output(target_data, output_path):
    """
    Step 4: Save the updated file.
    """
    print(f"\n{'='*60}")
    print("STEP 4: Saving updated file")
    print(f"{'='*60}")

    wb = target_data['workbook']
    wb.save(output_path)
    print(f"✓ Saved to: {output_path}")


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    source_file = sys.argv[1]
    target_file = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else target_file.replace('.xlsx', '_updated.xlsx')

    print(f"\n{'#'*60}")
    print("USDx Balances Weekly Update")
    print(f"{'#'*60}")
    print(f"Source file: {source_file}")
    print(f"Target file: {target_file}")
    print(f"Output file: {output_file}")

    try:
        # Step 1: Read source
        source_data = read_source_file(source_file)

        # Step 2: Read target
        target_data = read_target_file(target_file)

        # Step 3: Update
        stats = update_target_file(source_data, target_data)

        if stats is None:
            return

        # Step 4: Save
        save_output(target_data, output_file)

        # Step 5: Summary
        print(f"\n{'='*60}")
        print("STEP 5: SUMMARY")
        print(f"{'='*60}")
        print(f"Balance date: {source_data['balance_date'].strftime('%Y-%m-%d')}")
        print(f"Records matched: {stats['matched']}")
        print(f"Records missing from source: {stats['missing']}")
        print(f"New refs in source: {len(stats['new_refs'])}")
        print(f"\n✓ Update complete!")
        print(f"{'='*60}\n")

        # Close workbook
        target_data['workbook'].close()

    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
