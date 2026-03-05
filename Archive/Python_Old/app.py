"""
Balance Updater Web Application

A local web app to process USDx balance files and update Activity Rollforward files.
"""

from flask import Flask, render_template, request, send_file, jsonify
import os
import openpyxl
from datetime import datetime
from werkzeug.utils import secure_filename
import traceback

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def process_files(weekly_file_path, rollforward_file_path):
    """
    Main processing function:
    1. Read column K (Current Week) from USDx Balances tab (K10 header)
    2. Find last date in row 14 of Beginning Balances tab
    3. Paste column K data into the next column
    4. Copy formula from row 10 to generate new date header
    """

    log = []

    try:
        log.append("="*60)
        log.append("STEP 1: Reading Column K from USDx Balances")
        log.append("="*60)

        # Open source file
        wb_source = openpyxl.load_workbook(weekly_file_path, data_only=True)
        sheet_source = wb_source["USDx Balances"]

        # Column K is column 11, header in row 10
        balance_col = 11  # Column K
        ref_col = 2  # Column B for matching
        header_row_source = 10  # K10 contains "Current Week"
        data_start_row = 13

        balance_header = sheet_source.cell(header_row_source, balance_col).value
        log.append(f"Column K header (K10): {balance_header}")

        # Extract Current Week balances (column K) with reference numbers
        balances = {}
        for row in range(data_start_row, sheet_source.max_row + 1):
            ref_num = sheet_source.cell(row, ref_col).value
            balance = sheet_source.cell(row, balance_col).value

            if ref_num is not None:
                ref_str = str(ref_num).strip()
                # Skip section headers
                if not any(keyword in ref_str.lower() for keyword in ['bullish', 'coindesk', 'fiat', 'balance', 'total']):
                    try:
                        balance_val = float(balance) if balance is not None else 0.0
                    except (ValueError, TypeError):
                        balance_val = 0.0
                    balances[ref_str] = balance_val

        log.append(f"Extracted {len(balances)} Current Week balances from column K")

        wb_source.close()

        log.append("")
        log.append("="*60)
        log.append("STEP 2: Opening Activity Rollforward file")
        log.append("="*60)

        # Open target file (preserve formulas)
        wb_target = openpyxl.load_workbook(rollforward_file_path)

        # Find Beginning Balances sheet
        if "Beginning Balances" in wb_target.sheetnames:
            sheet_target = wb_target["Beginning Balances"]
            log.append("Found 'Beginning Balances' sheet")
        else:
            return {
                'success': False,
                'log': log,
                'error': "Beginning Balances sheet not found in Activity Rollforward file"
            }

        log.append("")
        log.append("="*60)
        log.append("STEP 3: Finding last date in row 14")
        log.append("="*60)

        # Find the last date in row 14
        last_col_with_date = 0
        for col in range(1, sheet_target.max_column + 1):
            cell_val = sheet_target.cell(14, col).value
            if cell_val is not None:
                last_col_with_date = col

        if last_col_with_date == 0:
            return {
                'success': False,
                'log': log,
                'error': "No dates found in row 14 of Beginning Balances"
            }

        log.append(f"Last date found in column {openpyxl.utils.get_column_letter(last_col_with_date)} (row 14)")

        # The new column is one to the right
        new_col = last_col_with_date + 1
        log.append(f"Will paste into column {openpyxl.utils.get_column_letter(new_col)}")

        log.append("")
        log.append("="*60)
        log.append("STEP 4: Copying formatting from previous column")
        log.append("="*60)

        # Copy formatting from previous column to new column
        from copy import copy
        source_col = last_col_with_date
        for row in range(1, sheet_target.max_row + 1):
            source_cell = sheet_target.cell(row, source_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy all formatting attributes
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        log.append(f"Copied formatting from column {openpyxl.utils.get_column_letter(source_col)}")

        log.append("")
        log.append("="*60)
        log.append("STEP 5: Copying formula from row 10 for date header")
        log.append("="*60)

        # Copy formula from row 10 (previous column) to generate new date
        formula_source_cell = sheet_target.cell(10, last_col_with_date)
        formula_target_cell = sheet_target.cell(10, new_col)

        # Check if source has a formula
        if formula_source_cell.value and isinstance(formula_source_cell.value, str) and formula_source_cell.value.startswith('='):
            # Copy the formula (openpyxl will adjust references automatically)
            formula_target_cell.value = formula_source_cell.value
            log.append(f"Copied formula from {openpyxl.utils.get_column_letter(last_col_with_date)}10 to {openpyxl.utils.get_column_letter(new_col)}10")
        else:
            # If no formula, just note it
            log.append(f"No formula found in row 10, column {openpyxl.utils.get_column_letter(last_col_with_date)}")
            log.append(f"Cell value: {formula_source_cell.value}")

        log.append("")
        log.append("="*60)
        log.append("STEP 6: Pasting Current Week balances")
        log.append("="*60)

        # Match and paste balances by Column B
        ref_col_target = 2  # Column B
        data_start_row_target = 15  # Data starts at row 15
        matched = 0
        not_matched = 0

        for row in range(data_start_row_target, sheet_target.max_row + 1):
            ref_num = sheet_target.cell(row, ref_col_target).value

            if ref_num is not None:
                ref_str = str(ref_num).strip()

                if ref_str in balances:
                    sheet_target.cell(row, new_col).value = balances[ref_str]
                    matched += 1
                else:
                    sheet_target.cell(row, new_col).value = 0.0
                    not_matched += 1

        log.append(f"✓ Matched and pasted: {matched} records")
        if not_matched > 0:
            log.append(f"⚠ Not found in source: {not_matched} records (set to 0.0)")

        log.append("")
        log.append("="*60)
        log.append("STEP 7: Copying formulas from previous week")
        log.append("="*60)

        # Copy formulas/cells from previous column (last_col_with_date) to new column (new_col)
        # for rows 12-20 and 148 to last row
        prev_col = last_col_with_date
        formulas_copied = 0

        # Copy rows 12-20
        for row in range(12, 21):  # 12 to 20 inclusive
            source_cell = sheet_target.cell(row, prev_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy formula or value
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                target_cell.value = source_cell.value
                formulas_copied += 1
            else:
                target_cell.value = source_cell.value

        log.append(f"Copied cells for rows 12-20 from column {openpyxl.utils.get_column_letter(prev_col)}")

        # Find last row with data
        last_data_row = sheet_target.max_row

        # Copy rows 148 to last row
        for row in range(148, last_data_row + 1):
            source_cell = sheet_target.cell(row, prev_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy formula or value
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                target_cell.value = source_cell.value
                formulas_copied += 1
            else:
                target_cell.value = source_cell.value

        log.append(f"Copied cells for rows 148-{last_data_row} from column {openpyxl.utils.get_column_letter(prev_col)}")
        log.append(f"Total formulas copied: {formulas_copied}")

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Activity_Rollforward_Updated_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

        log.append("")
        log.append("="*60)
        log.append("STEP 8: Saving output file")
        log.append("="*60)

        wb_target.save(output_path)
        log.append(f"✓ Saved to: {output_filename}")

        wb_target.close()

        log.append("")
        log.append("="*60)
        log.append("SUMMARY")
        log.append("="*60)
        log.append(f"Column pasted to: {openpyxl.utils.get_column_letter(new_col)}")
        log.append(f"Current Week balances matched: {matched}")
        log.append(f"Formulas copied: {formulas_copied}")
        log.append(f"✓ Process completed successfully!")
        log.append("="*60)

        return {
            'success': True,
            'log': log,
            'output_file': output_filename,
            'stats': {
                'column': openpyxl.utils.get_column_letter(new_col),
                'total_balances': len(balances),
                'balances_matched': matched,
                'formulas_copied': formulas_copied
            }
        }

    except Exception as e:
        log.append(f"\n✗ ERROR: {str(e)}")
        log.append(traceback.format_exc())
        return {
            'success': False,
            'log': log,
            'error': str(e)
        }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        # Check if files are present
        if 'weekly_file' not in request.files or 'rollforward_file' not in request.files:
            return jsonify({'success': False, 'error': 'Both files are required'})

        weekly_file = request.files['weekly_file']
        rollforward_file = request.files['rollforward_file']

        # Check if files are selected
        if weekly_file.filename == '' or rollforward_file.filename == '':
            return jsonify({'success': False, 'error': 'Please select both files'})

        # Validate file types
        if not (allowed_file(weekly_file.filename) and allowed_file(rollforward_file.filename)):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed'})

        # Save uploaded files
        weekly_filename = secure_filename(weekly_file.filename)
        rollforward_filename = secure_filename(rollforward_file.filename)

        weekly_path = os.path.join(app.config['UPLOAD_FOLDER'], weekly_filename)
        rollforward_path = os.path.join(app.config['UPLOAD_FOLDER'], rollforward_filename)

        weekly_file.save(weekly_path)
        rollforward_file.save(rollforward_path)

        # Process the files
        result = process_files(weekly_path, rollforward_path)

        # Clean up uploaded files
        os.remove(weekly_path)
        os.remove(rollforward_path)

        return jsonify(result)

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'log': [traceback.format_exc()]
        })


@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@app.route('/outputs')
def list_outputs():
    """List all output files"""
    try:
        files = []
        for filename in os.listdir(app.config['OUTPUT_FOLDER']):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
                files.append({
                    'name': filename,
                    'size': os.path.getsize(file_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
        files.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify(files)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Create folders if they don't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

    print("\n" + "="*60)
    print("Balance Updater Web App")
    print("="*60)
    print("\nStarting server...")
    print("Open your browser and go to: http://localhost:5000")
    print("\nPress CTRL+C to stop the server")
    print("="*60 + "\n")

    app.run(debug=True, host='0.0.0.0', port=5000)
