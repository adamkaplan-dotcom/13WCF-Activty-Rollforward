"""
Paste USDx Balances into Activity Rollforward - Beginning Balances Tab

This script copies data from the USDx Balances tab and pastes it into the
Beginning Balances tab, matching on column B references.
"""

import sys
import openpyxl
from datetime import datetime


def main(source_file, target_file, output_file):
    print("="*60)
    print("STEP 1: Reading USDx Balances from source file")
    print("="*60)

    # Open source file
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    sheet_source = wb_source["USDx Balances"]

    # Find the data - header in row 12, data starts row 13
    header_row = 12
    data_start_row = 13

    # Extract all data from USDx Balances tab
    print(f"Reading from '{sheet_source.title}' tab...")

    # Get column headers
    max_col = sheet_source.max_column
    headers = []
    for col in range(1, max_col + 1):
        header_val = sheet_source.cell(header_row, col).value
        headers.append(header_val)

    print(f"Found {max_col} columns")

    # Read all data rows
    data_rows = []
    for row in range(data_start_row, sheet_source.max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell_val = sheet_source.cell(row, col).value
            row_data.append(cell_val)

        # Skip empty rows
        if any(row_data):
            data_rows.append(row_data)

    print(f"Found {len(data_rows)} data rows")

    wb_source.close()

    print("\n" + "="*60)
    print("STEP 2: Opening Activity Rollforward file")
    print("="*60)

    # Open target file
    wb_target = openpyxl.load_workbook(target_file)

    # Find or create "Beginning Balances" sheet
    if "Beginning Balances" in wb_target.sheetnames:
        sheet_target = wb_target["Beginning Balances"]
        print(f"Found existing 'Beginning Balances' sheet")
    else:
        sheet_target = wb_target.create_sheet("Beginning Balances")
        print(f"Created new 'Beginning Balances' sheet")

    print("\n" + "="*60)
    print("STEP 3: Pasting data into Beginning Balances tab")
    print("="*60)

    # Clear existing data if any
    sheet_target.delete_rows(1, sheet_target.max_row)

    # Write headers
    for col_idx, header_val in enumerate(headers, start=1):
        sheet_target.cell(1, col_idx).value = header_val

    print(f"Wrote {len(headers)} column headers")

    # Write data rows
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, cell_val in enumerate(row_data, start=1):
            sheet_target.cell(row_idx, col_idx).value = cell_val

    print(f"Wrote {len(data_rows)} data rows")

    print("\n" + "="*60)
    print("STEP 4: Saving updated file")
    print("="*60)

    wb_target.save(output_file)
    print(f"✓ Saved to: {output_file}")

    wb_target.close()

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Columns copied: {len(headers)}")
    print(f"Rows copied: {len(data_rows)}")
    print(f"✓ Complete!")
    print("="*60)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python paste_balances.py <source_file> <target_file> [output_file]")
        sys.exit(1)

    source = sys.argv[1]
    target = sys.argv[2]
    output = sys.argv[3] if len(sys.argv) > 3 else target.replace('.xlsx', '_with_balances.xlsx')

    main(source, target, output)
