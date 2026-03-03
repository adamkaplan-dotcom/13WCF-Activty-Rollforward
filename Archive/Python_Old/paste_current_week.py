"""
Paste Current Week Balances to the Rightmost Empty Column

This script:
1. Reads Current Week balances from column K of USDx Balances tab
2. Finds the first empty column on the right in the target file
3. Pastes the balances matching on column B references
"""

import sys
import openpyxl
from datetime import datetime


def main(source_file, target_file, output_file):
    print("="*60)
    print("STEP 1: Reading Current Week balances from source file")
    print("="*60)

    # Open source file
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    sheet_source = wb_source["USDx Balances"]

    # Header row is 12, data starts at row 13
    header_row = 12
    data_start_row = 13
    ref_col = 2  # Column B
    balance_col = 11  # Column K

    # Get the date/header for column K
    balance_date = sheet_source.cell(header_row, balance_col).value
    print(f"Balance date from column K: {balance_date}")

    # Read all balances with their ref numbers
    balances = {}
    for row in range(data_start_row, sheet_source.max_row + 1):
        ref_num = sheet_source.cell(row, ref_col).value
        balance = sheet_source.cell(row, balance_col).value

        if ref_num is not None:
            # Skip section headers
            ref_str = str(ref_num).strip()
            if not any(keyword in ref_str.lower() for keyword in ['bullish', 'coindesk', 'fiat', 'balance', 'total']):
                # Convert balance to float
                try:
                    balance_val = float(balance) if balance is not None else 0.0
                except (ValueError, TypeError):
                    balance_val = 0.0

                balances[ref_str] = balance_val

    print(f"Extracted {len(balances)} balance records from column K")

    wb_source.close()

    print("\n" + "="*60)
    print("STEP 2: Opening target file and finding empty column")
    print("="*60)

    # Open target file
    wb_target = openpyxl.load_workbook(target_file)

    # Use Beginning Balances sheet
    if "Beginning Balances" in wb_target.sheetnames:
        sheet_target = wb_target["Beginning Balances"]
        print(f"Using 'Beginning Balances' sheet")
    else:
        print("ERROR: 'Beginning Balances' sheet not found")
        wb_target.close()
        return

    # Find the first empty column (rightmost)
    header_row_target = 1
    last_col = 1

    for col in range(1, sheet_target.max_column + 2):
        header_val = sheet_target.cell(header_row_target, col).value
        if header_val is None:
            last_col = col
            break
        last_col = col + 1

    print(f"First empty column: {openpyxl.utils.get_column_letter(last_col)}")

    print("\n" + "="*60)
    print("STEP 3: Matching and pasting balances")
    print("="*60)

    # Write header in the new column
    sheet_target.cell(header_row_target, last_col).value = balance_date
    print(f"Added header: {balance_date}")

    # Read ref numbers from target file and match
    ref_col_target = 2  # Column B
    data_start_row_target = 2
    matched = 0
    not_matched = 0

    for row in range(data_start_row_target, sheet_target.max_row + 1):
        ref_num = sheet_target.cell(row, ref_col_target).value

        if ref_num is not None:
            ref_str = str(ref_num).strip()

            if ref_str in balances:
                sheet_target.cell(row, last_col).value = balances[ref_str]
                matched += 1
            else:
                sheet_target.cell(row, last_col).value = 0.0
                not_matched += 1

    print(f"✓ Matched and pasted: {matched} records")
    if not_matched > 0:
        print(f"⚠ Not found in source: {not_matched} records (set to 0.0)")

    print("\n" + "="*60)
    print("STEP 4: Saving updated file")
    print("="*60)

    wb_target.save(output_file)
    print(f"✓ Saved to: {output_file}")

    wb_target.close()

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Balance date: {balance_date}")
    print(f"Column added: {openpyxl.utils.get_column_letter(last_col)}")
    print(f"Records matched: {matched}")
    print(f"Records not matched: {not_matched}")
    print(f"✓ Complete!")
    print("="*60)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python paste_current_week.py <source_file> <target_file> [output_file]")
        sys.exit(1)

    source = sys.argv[1]
    target = sys.argv[2]
    output = sys.argv[3] if len(sys.argv) > 3 else target.replace('.xlsx', '_updated.xlsx')

    main(source, target, output)
