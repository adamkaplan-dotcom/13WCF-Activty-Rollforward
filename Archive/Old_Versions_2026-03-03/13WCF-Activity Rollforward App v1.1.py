"""
Balance Updater Web Application

A local web app to process USDx balance files and update Activity Rollforward files.
"""

from flask import Flask, render_template, request, send_file, jsonify
import os
import openpyxl
from datetime import datetime
from werkzeug.utils import secure_filename
import traceback
import re

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def adjust_formula_columns(formula, column_offset=1):
    """
    Adjust column references in an Excel formula by a given offset.

    Example:
        Input: "=CC14+7", offset=1
        Output: "=CD14+7"

        Input: "=INDEX($B:$B,MATCH(CC14,$A:$A,0))", offset=1
        Output: "=INDEX($B:$B,MATCH(CD14,$A:$A,0))"

    Args:
        formula: Excel formula string starting with '='
        column_offset: Number of columns to shift right (default 1)

    Returns:
        Adjusted formula string
    """
    if not formula or not formula.startswith('='):
        return formula

    def col_to_num(col_letters):
        """Convert column letters to number (A=1, Z=26, AA=27, etc.)"""
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        return col_num

    def num_to_col(num):
        """Convert number to column letters (1=A, 26=Z, 27=AA, etc.)"""
        col_letters = ''
        while num > 0:
            num -= 1
            col_letters = chr(ord('A') + (num % 26)) + col_letters
            num //= 26
        return col_letters

    def adjust_cell_ref(match):
        """Adjust a single cell reference"""
        prefix_char = match.group(1) or ''
        dollar1 = match.group(2) or ''
        col_letters = match.group(3)
        dollar2 = match.group(4) or ''
        row_number = match.group(5) or ''

        # Only adjust if column reference is relative (no $ before column)
        if dollar1:
            # Absolute column reference - don't adjust
            return prefix_char + dollar1 + col_letters + dollar2 + row_number
        else:
            # Relative column reference - adjust it
            col_num = col_to_num(col_letters)
            new_col = num_to_col(col_num + column_offset)
            return prefix_char + dollar1 + new_col + dollar2 + row_number

    # Pattern matches cell references with row numbers (not bare column letters)
    # This avoids matching function names like INDEX, MATCH, etc.
    # Matches: A1, $A1, A$1, $A$1, CC14, $CC$14, etc.
    # Also matches ranges like CC14:CC20
    pattern = r'(^|[^A-Za-z])(\$?)([A-Z]+)(\$?)(\d+)'

    # First pass: adjust cell references with row numbers
    adjusted_formula = re.sub(pattern, adjust_cell_ref, formula)

    # Second pass: adjust column-only ranges like $B:$B or A:Z
    def adjust_col_range(match):
        prefix = match.group(1) or ''
        dollar1 = match.group(2) or ''
        col1 = match.group(3)
        colon = match.group(4)
        dollar2 = match.group(5) or ''
        col2 = match.group(6)
        suffix = match.group(7) or ''

        # Adjust first column if relative
        if dollar1:
            new_col1 = col1
        else:
            new_col1 = num_to_col(col_to_num(col1) + column_offset)

        # Adjust second column if relative
        if dollar2:
            new_col2 = col2
        else:
            new_col2 = num_to_col(col_to_num(col2) + column_offset)

        return prefix + dollar1 + new_col1 + colon + dollar2 + new_col2 + suffix

    # Match column ranges like $B:$B, A:Z, etc.
    range_pattern = r'([^A-Za-z])(\$?)([A-Z]+)(:)(\$?)([A-Z]+)([^A-Za-z]|$)'
    adjusted_formula = re.sub(range_pattern, adjust_col_range, adjusted_formula)

    return adjusted_formula

def process_files(weekly_file_path, rollforward_file_path, bth_file_path=None):
    """
    Main processing function:
    1. Read column K (Current Week) from USDx Balances tab (K10 header)
    2. Find last date in row 14 of Beginning Balances tab
    3. Paste column K data into the next column
    4. Copy formula from row 10 to generate new date header
    5. Process Cockpit tab (copy formulas from last column to next)
    6. If BTH file provided, match date and paste Total BTH Financing to Cockpit row 20
    """

    log = []

    try:
        log.append("="*60)
        log.append("STEP 1: Reading Column K from USDx Balances")
        log.append("="*60)

        # Open source file
        wb_source = openpyxl.load_workbook(weekly_file_path, data_only=True)
        sheet_source = wb_source["USDx Balances"]

        # Column K is column 11, header in row 10
        balance_col = 11  # Column K
        ref_col = 2  # Column B for matching
        header_row_source = 10  # K10 contains "Current Week"
        data_start_row = 13

        balance_header = sheet_source.cell(header_row_source, balance_col).value
        log.append(f"Column K header (K10): {balance_header}")

        # Extract Current Week balances (column K) with reference numbers
        balances = {}
        for row in range(data_start_row, sheet_source.max_row + 1):
            ref_num = sheet_source.cell(row, ref_col).value
            balance = sheet_source.cell(row, balance_col).value

            if ref_num is not None:
                ref_str = str(ref_num).strip()
                # Skip section headers
                if not any(keyword in ref_str.lower() for keyword in ['bullish', 'coindesk', 'fiat', 'balance', 'total']):
                    try:
                        balance_val = float(balance) if balance is not None else 0.0
                    except (ValueError, TypeError):
                        balance_val = 0.0
                    balances[ref_str] = balance_val

        log.append(f"Extracted {len(balances)} Current Week balances from column K")

        wb_source.close()

        log.append("")
        log.append("="*60)
        log.append("STEP 2: Opening Activity Rollforward file")
        log.append("="*60)

        # Open target file (preserve formulas)
        wb_target = openpyxl.load_workbook(rollforward_file_path)

        # Find Beginning Balances sheet
        if "Beginning Balances" in wb_target.sheetnames:
            sheet_target = wb_target["Beginning Balances"]
            log.append("Found 'Beginning Balances' sheet")
        else:
            return {
                'success': False,
                'log': log,
                'error': "Beginning Balances sheet not found in Activity Rollforward file"
            }

        log.append("")
        log.append("="*60)
        log.append("STEP 3: Finding last date in row 14")
        log.append("="*60)

        # Find the last date in row 14
        last_col_with_date = 0
        for col in range(1, sheet_target.max_column + 1):
            cell_val = sheet_target.cell(14, col).value
            if cell_val is not None:
                last_col_with_date = col

        if last_col_with_date == 0:
            return {
                'success': False,
                'log': log,
                'error': "No dates found in row 14 of Beginning Balances"
            }

        log.append(f"Last date found in column {openpyxl.utils.get_column_letter(last_col_with_date)} (row 14)")

        # The new column is one to the right
        new_col = last_col_with_date + 1
        log.append(f"Will paste into column {openpyxl.utils.get_column_letter(new_col)}")

        log.append("")
        log.append("="*60)
        log.append("STEP 4: Copying formatting from previous column")
        log.append("="*60)

        # Copy formatting from previous column to new column
        from copy import copy
        source_col = last_col_with_date
        for row in range(1, sheet_target.max_row + 1):
            source_cell = sheet_target.cell(row, source_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy all formatting attributes
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        log.append(f"Copied formatting from column {openpyxl.utils.get_column_letter(source_col)}")

        log.append("")
        log.append("="*60)
        log.append("STEP 5: Copying formula from row 10 for date header")
        log.append("="*60)

        # Copy formula from row 10 (previous column) to generate new date
        formula_source_cell = sheet_target.cell(10, last_col_with_date)
        formula_target_cell = sheet_target.cell(10, new_col)

        # Check if source has a formula
        if formula_source_cell.value and isinstance(formula_source_cell.value, str) and formula_source_cell.value.startswith('='):
            # Adjust column references in formula
            adjusted_formula = adjust_formula_columns(formula_source_cell.value, column_offset=1)
            formula_target_cell.value = adjusted_formula
            log.append(f"Copied formula from {openpyxl.utils.get_column_letter(last_col_with_date)}10 to {openpyxl.utils.get_column_letter(new_col)}10")
            log.append(f"  Original: {formula_source_cell.value}")
            log.append(f"  Adjusted: {adjusted_formula}")
        else:
            # If no formula, just note it
            log.append(f"No formula found in row 10, column {openpyxl.utils.get_column_letter(last_col_with_date)}")
            log.append(f"Cell value: {formula_source_cell.value}")

        log.append("")
        log.append("="*60)
        log.append("STEP 6: Pasting Current Week balances")
        log.append("="*60)

        # Match and paste balances by Column B
        ref_col_target = 2  # Column B
        data_start_row_target = 15  # Data starts at row 15
        matched = 0
        not_matched = 0
        copied_from_prev = 0

        # Special references that should copy from previous week instead of Weekly Balances file
        copy_from_previous_refs = ['11', '60', '76', '91']

        for row in range(data_start_row_target, sheet_target.max_row + 1):
            ref_num = sheet_target.cell(row, ref_col_target).value

            if ref_num is not None:
                ref_str = str(ref_num).strip()

                # Check if this is a special reference that should copy from previous week
                # Only applies to rows between 15 and 148
                if ref_str in copy_from_previous_refs and row <= 148:
                    # Copy value from previous week's column
                    prev_value = sheet_target.cell(row, last_col_with_date).value
                    sheet_target.cell(row, new_col).value = prev_value
                    copied_from_prev += 1
                    log.append(f"  Row {row} (Ref {ref_str}): Copied {prev_value} from previous week")
                elif ref_str in balances:
                    sheet_target.cell(row, new_col).value = balances[ref_str]
                    matched += 1
                else:
                    sheet_target.cell(row, new_col).value = 0.0
                    not_matched += 1

        log.append(f"✓ Matched and pasted: {matched} records")
        log.append(f"✓ Copied from previous week: {copied_from_prev} records (refs: 11, 60, 76, 91)")
        if not_matched > 0:
            log.append(f"⚠ Not found in source: {not_matched} records (set to 0.0)")

        log.append("")
        log.append("="*60)
        log.append("STEP 7: Copying and adjusting formulas from previous week")
        log.append("="*60)

        # Copy formulas/cells from previous column (last_col_with_date) to new column (new_col)
        # Adjust column references so formulas reference the correct columns
        prev_col = last_col_with_date
        formulas_copied = 0
        formulas_adjusted = 0

        # Copy rows 12-20
        log.append(f"Copying rows 12-20 from column {openpyxl.utils.get_column_letter(prev_col)} to {openpyxl.utils.get_column_letter(new_col)}...")
        for row in range(12, 21):  # 12 to 20 inclusive
            source_cell = sheet_target.cell(row, prev_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy formula or value
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                # Adjust column references in formula
                adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=1)
                target_cell.value = adjusted_formula
                formulas_copied += 1
                if adjusted_formula != source_cell.value:
                    formulas_adjusted += 1
            else:
                target_cell.value = source_cell.value

        log.append(f"✓ Copied {formulas_copied} formulas for rows 12-20")

        # Copy rows 148-239 (as specified by user)
        log.append(f"Copying rows 148-239 from column {openpyxl.utils.get_column_letter(prev_col)} to {openpyxl.utils.get_column_letter(new_col)}...")
        formulas_copied_148 = 0
        for row in range(148, 240):  # 148 to 239 inclusive
            source_cell = sheet_target.cell(row, prev_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy formula or value
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                # Adjust column references in formula
                adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=1)
                target_cell.value = adjusted_formula
                formulas_copied_148 += 1
                if adjusted_formula != source_cell.value:
                    formulas_adjusted += 1
            else:
                target_cell.value = source_cell.value

        log.append(f"✓ Copied {formulas_copied_148} formulas for rows 148-239")
        log.append(f"✓ Total formulas copied: {formulas_copied + formulas_copied_148}")
        log.append(f"✓ Total formulas adjusted: {formulas_adjusted}")

        log.append("")
        log.append("="*60)
        log.append("STEP 8: Processing Cockpit Tab")
        log.append("="*60)

        # Check if Cockpit tab exists
        if "Cockpit" in wb_target.sheetnames:
            sheet_cockpit = wb_target["Cockpit"]
            log.append("✓ Found Cockpit tab")

            # Use the SAME column we just created in Beginning Balances
            # This keeps both tabs in sync
            cockpit_prev_col = new_col - 1  # Previous column (e.g., BJ if new_col is BK)
            cockpit_new_col = new_col        # Same as Beginning Balances (e.g., BK)

            log.append(f"Copying from column {openpyxl.utils.get_column_letter(cockpit_prev_col)} to {openpyxl.utils.get_column_letter(cockpit_new_col)}")

            # Copy formatting from previous column
            from copy import copy
            for row in range(1, sheet_cockpit.max_row + 1):
                source_cell = sheet_cockpit.cell(row, cockpit_prev_col)
                target_cell = sheet_cockpit.cell(row, cockpit_new_col)

                # Copy formatting
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

            # Copy formulas/values from previous column to new column
            cockpit_formulas_copied = 0
            for row in range(1, sheet_cockpit.max_row + 1):
                source_cell = sheet_cockpit.cell(row, cockpit_prev_col)
                target_cell = sheet_cockpit.cell(row, cockpit_new_col)

                # Copy formula or value
                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    # Adjust column references in formula
                    adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=1)
                    target_cell.value = adjusted_formula
                    cockpit_formulas_copied += 1
                else:
                    target_cell.value = source_cell.value

            log.append(f"✓ Copied {cockpit_formulas_copied} formulas to Cockpit tab")

                # Process BTH file if provided
                if bth_file_path:
                    log.append("")
                    log.append("="*60)
                    log.append("STEP 9: Processing BTH Investments File")
                    log.append("="*60)

                    try:
                        wb_bth = openpyxl.load_workbook(bth_file_path, data_only=True)

                        if "Summary" in wb_bth.sheetnames:
                            sheet_bth = wb_bth["Summary"]
                            log.append("✓ Found Summary tab in BTH file")

                            # Find the date in row 14 of Beginning Balances (which we just added to)
                            target_date = sheet_target.cell(14, new_col).value
                            log.append(f"Looking for date: {target_date}")

                            # Search for matching date in row 30 of BTH Summary tab
                            matched_bth_col = None
                            for col in range(1, sheet_bth.max_column + 1):
                                bth_date = sheet_bth.cell(30, col).value
                                if bth_date and isinstance(bth_date, datetime):
                                    # Compare dates (ignore time component)
                                    if isinstance(target_date, datetime):
                                        if bth_date.date() == target_date.date():
                                            matched_bth_col = col
                                            break

                            if matched_bth_col:
                                # Get Total BTH Financing from row 42
                                bth_financing_total = sheet_bth.cell(42, matched_bth_col).value
                                log.append(f"✓ Matched date in column {openpyxl.utils.get_column_letter(matched_bth_col)}")
                                log.append(f"Total BTH Financing (row 42): {bth_financing_total}")

                                # Paste to Cockpit tab, row 20, new column
                                sheet_cockpit.cell(20, cockpit_new_col).value = bth_financing_total
                                log.append(f"✓ Pasted {bth_financing_total} to Cockpit {openpyxl.utils.get_column_letter(cockpit_new_col)}20")
                            else:
                                log.append("⚠ Could not find matching date in BTH file")

                        wb_bth.close()

                    except Exception as e:
                        log.append(f"⚠ Error processing BTH file: {str(e)}")
                        log.append(traceback.format_exc())
        else:
            log.append("⚠ Cockpit tab not found in Activity Rollforward file")

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Activity_Rollforward_Updated_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

        log.append("")
        log.append("="*60)
        log.append("STEP 10: Saving output file")
        log.append("="*60)

        wb_target.save(output_path)
        log.append(f"✓ Saved to: {output_filename}")

        wb_target.close()

        log.append("")
        log.append("="*60)
        log.append("SUMMARY")
        log.append("="*60)
        log.append(f"Column pasted to: {openpyxl.utils.get_column_letter(new_col)}")
        log.append(f"Current Week balances matched: {matched}")
        log.append(f"Formulas copied and adjusted: {formulas_copied + formulas_copied_148}")
        log.append(f"✓ Process completed successfully!")
        log.append("="*60)

        return {
            'success': True,
            'log': log,
            'output_file': output_filename,
            'stats': {
                'column': openpyxl.utils.get_column_letter(new_col),
                'total_balances': len(balances),
                'balances_matched': matched,
                'formulas_copied': formulas_copied + formulas_copied_148
            }
        }

    except Exception as e:
        log.append(f"\n✗ ERROR: {str(e)}")
        log.append(traceback.format_exc())
        return {
            'success': False,
            'log': log,
            'error': str(e)
        }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        # Check if required files are present
        if 'weekly_file' not in request.files or 'rollforward_file' not in request.files:
            return jsonify({'success': False, 'error': 'Weekly Balances and Activity Rollforward files are required'})

        weekly_file = request.files['weekly_file']
        rollforward_file = request.files['rollforward_file']
        bth_file = request.files.get('bth_file')  # Optional

        # Check if required files are selected
        if weekly_file.filename == '' or rollforward_file.filename == '':
            return jsonify({'success': False, 'error': 'Please select Weekly Balances and Activity Rollforward files'})

        # Validate file types
        if not (allowed_file(weekly_file.filename) and allowed_file(rollforward_file.filename)):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed'})

        if bth_file and bth_file.filename != '' and not allowed_file(bth_file.filename):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed for BTH file'})

        # Save uploaded files
        weekly_filename = secure_filename(weekly_file.filename)
        rollforward_filename = secure_filename(rollforward_file.filename)

        weekly_path = os.path.join(app.config['UPLOAD_FOLDER'], weekly_filename)
        rollforward_path = os.path.join(app.config['UPLOAD_FOLDER'], rollforward_filename)

        weekly_file.save(weekly_path)
        rollforward_file.save(rollforward_path)

        # Save BTH file if provided
        bth_path = None
        if bth_file and bth_file.filename != '':
            bth_filename = secure_filename(bth_file.filename)
            bth_path = os.path.join(app.config['UPLOAD_FOLDER'], bth_filename)
            bth_file.save(bth_path)

        # Process the files
        result = process_files(weekly_path, rollforward_path, bth_path)

        # Clean up uploaded files
        os.remove(weekly_path)
        os.remove(rollforward_path)
        if bth_path and os.path.exists(bth_path):
            os.remove(bth_path)

        return jsonify(result)

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'log': [traceback.format_exc()]
        })


@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@app.route('/outputs')
def list_outputs():
    """List all output files"""
    try:
        files = []
        for filename in os.listdir(app.config['OUTPUT_FOLDER']):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
                files.append({
                    'name': filename,
                    'size': os.path.getsize(file_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
        files.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify(files)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Create folders if they don't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

    print("\n" + "="*60)
    print("Balance Updater Web App")
    print("="*60)
    print("\nStarting server...")
    print("Open your browser and go to: http://localhost:5000")
    print("\nPress CTRL+C to stop the server")
    print("="*60 + "\n")

    app.run(debug=True, host='0.0.0.0', port=5000)
